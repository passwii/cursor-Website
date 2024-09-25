from flask import Blueprint, render_template, request, send_file, redirect, flash, url_for
import os
import io
import datetime
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, NamedStyle
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter

main_monthly = Blueprint('main_monthly', __name__)

def reset_style(filename):
    wb = load_workbook(filename)
    line_t = Side(style='thin', color='000000')

    sty1 = NamedStyle(name='sty1')
    sty1.font = Font(name='等线', size=12, b=True)
    sty1.alignment = Alignment(horizontal='left', vertical='center')
    sty1.border = Border(top=line_t, bottom=line_t, left=line_t, right=line_t)

    sty2 = NamedStyle(name='sty2')
    sty2.font = Font(name='等线', size=12)
    sty2.alignment = Alignment(horizontal='left', vertical='center')
    sty2.border = Border(top=line_t, bottom=line_t, left=line_t, right=line_t)

    sty3 = NamedStyle(name='sty3')
    sty3.font = Font(name='等线', size=12)
    sty3.alignment = Alignment(horizontal='center', vertical='center')
    sty3.border = Border(top=line_t, bottom=line_t, left=line_t, right=line_t)

    for ws1 in wb:
        nrows = ws1.max_row
        ncols = ws1.max_column
        width = 12
        height = 15
        for r in range(1, nrows + 1):
            for c in range(1, ncols + 1):
                if r == 1:
                    ws1.cell(r, c).style = sty1
                    ws1.row_dimensions[r].height = height
                else:
                    ws1.cell(r, c).style = sty2
                    ws1.column_dimensions[get_column_letter(
                        c)].width = width
    wb['报表总览'].sheet_properties.tabColor = "1072BA"
    return wb.save(filename)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'csv'}

def save_file(file, folder, filename):
    if not os.path.exists(folder):
        os.makedirs(folder)
    filepath = os.path.join(folder, filename)
    file.save(filepath)
    return filepath

def process_monthly_report(project_name, report_date, payment_range_report):
    current_time = datetime.datetime.now().strftime('%H-%M-%S')
    source_folder = os.getcwd()
    os.chdir(source_folder)

    report_name = f'{project_name} 美国站 {report_date}'

    project_folder_path = os.path.join(source_folder, 'project', project_name, '月报')
    os.makedirs(project_folder_path, exist_ok=True)

    tmp_folder_path = os.path.join(source_folder, 'project', project_name, 'tmp')
    os.makedirs(tmp_folder_path, exist_ok=True)

    payment_range_report_path = f'{project_name}_payment_range_report_{report_date}.csv'

    with open(payment_range_report_path, 'wb') as f:
        f.write(payment_range_report.read())

    payment_range_report = pd.read_csv(payment_range_report_path, thousands=',', skiprows=7, encoding='utf-8')

    files_to_copy = [
        (payment_range_report_path, f'{tmp_folder_path}/{project_name}_payment_range_report_{report_date}_{current_time}.csv')
    ]

    for src, dst in files_to_copy:
        shutil.copy(src, dst)
        os.remove(src)
    PRR = payment_range_report
    PRR.fillna(0, inplace=True)
    PRR['quantity'] = PRR['quantity'].astype(int)

    # 订单销售额
    Order = PRR.loc[PRR['type'].isin(['Order'])]
    Liquidation = PRR.loc[PRR['type'].isin(['Liquidations'])]
    FBA_product_sales = round(Order['product sales'].sum(), 2)
    Liquidation_product_sales = round(Liquidation['product sales'].sum(), 2)
    Liquidation_processing_fee = round(Liquidation['other transaction fees'].sum(), 2)
    Shipping_credits = round(Order['shipping credits'].sum(), 2)
    Promotional_rebates = round(Order['promotional rebates'].sum(), 2)
    Gift_wrap_credits = round(Order['gift wrap credits'].sum(), 2)

    # FBA & FBM 平台费
    FBA_selling_fees = round(Order['selling fees'].sum(), 2)
    FBA_transaction_fees = round(Order['fba fees'].sum(), 2)
    # 销售SKU明细
    skuGroup = Order.groupby(['sku'], as_index=False).agg({'quantity': 'sum'})
    Order_QTY = int(Order['quantity'].sum())

    # Retrocharge 计算
    Order_Retrocharge = PRR.loc[PRR['type'].isin(['Order Retrocharge'])]
    Order_Retrocharge_Base_Tax = Order_Retrocharge.loc[
        Order_Retrocharge['description'].isin(['Base Tax'])]
    Order_Retrocharge_Base_Tax = round(Order_Retrocharge_Base_Tax['total'].sum(), 2)
    Order_Retrocharge_Shipping_Tax = round(
        Order_Retrocharge.loc[Order_Retrocharge['description'].isin(['Shipping Tax'])][
            'total'].sum(), 2)
    Order_Retrocharge_MarketplaceFacilitator = round(
        Order_Retrocharge.loc[
            Order_Retrocharge['description'].isin(['MarketplaceFacilitator'])]['total'].sum(), 2)

    # Refund 计算
    Refund = PRR.loc[PRR['type'].isin(['Refund'])]
    Refund_other = round(Refund['other'].sum(), 2)
    FBA_product_sale_refunds = round(Refund['product sales'].sum(), 2)
    Shipping_credits_refunds = round(Refund['shipping credits'].sum(), 2)
    Promotional_rebates_refunds = round(Refund['promotional rebates'].sum(), 2)
    Gift_wrap_credits_refunds = round(Refund['gift wrap credits'].sum(), 2)

    FBA_selling_fees_refunds = round(Refund['selling fees'].sum(), 2)
    FBA_transaction_fees_refunds = round(Refund['fba fees'].sum(), 2)
    refund_skuGroup = Refund.groupby(['sku'], as_index=False).agg({'quantity': 'sum'})
    Refund_QTY = int(Refund['quantity'].sum())

    # Adjustment 赔偿等
    Adjustment = PRR.loc[PRR['type'].isin(['Adjustment'])]

    # Fee Adjustment 亚马逊 返还 的派送费(产品发货尺寸, 重新测量的费用) 
    Fee_Adjustment = PRR.loc[PRR['type'].isin(['Fee Adjustment'])]
    Fee_Adjustment_FBA_Fee_Adjustment = round(
        Fee_Adjustment.loc[
            Fee_Adjustment['description'].isin(['FBA Weight/Dimension Change'])][
                'total'].sum(), 2)

    FBA_Inventory_Reimbursement = round(Adjustment['total'].sum(), 2)

    # Service Fee
    Service_Fee = PRR.loc[PRR['type'].isin(['Service Fee'])]
    Cost_of_Advertising_Selection = Service_Fee.loc[Service_Fee['description'].isin(['Cost of Advertising'])]
    Cost_of_Advertising = round(Cost_of_Advertising_Selection['total'].sum(), 2)
    Subscription = round(
        Service_Fee.loc[Service_Fee['description'].isin(['Subscription'])]['total'].sum(), 2)
    Subscription_Selection = Service_Fee.loc[Service_Fee['description'].isin(['Subscription'])]
    AGL_Selection = Service_Fee.loc[Service_Fee['description'].isin(['FBA International Freight', 'FBA International Freight Duties and Taxes Charge'])]

    # LD 等费用
    LD = PRR.loc[PRR['type'].isin(['Deal Fee'])]
    Lightning_Deal_Fee = LD['total'].sum()

    # FBA Inventory Fee 库存费用
    FBA_Inventory_Fee = PRR.loc[PRR['type'].isin(['FBA Inventory Fee'])]
    All_Inventory_Fee = round(FBA_Inventory_Fee['total'].sum(), 2)

    # 销售额 = sales + 促销返点 + 赔偿 +所有税费和预扣
    Order_sales = round(FBA_product_sales + Shipping_credits + Promotional_rebates, 2)

    # 退款
    Refunds = round(
        FBA_product_sale_refunds + Shipping_credits_refunds + Promotional_rebates_refunds + Refund_other, 2)

    # 平台处理费
    AMZ = round(FBA_selling_fees, 2)
    FBA = round(FBA_transaction_fees, 2)
    AMZnFBA = round(AMZ + FBA, 2)
    AMZnFBA_Selection = PRR.loc[PRR['type'].isin(['Order', 'Refund'])]  # Order 亚马逊订单, 不计算退款，退款在Refund中和 PDF 中单独计算

    # 平台推广费用
    spFee = Cost_of_Advertising
    spCount = format(abs(spFee) / Order_sales, '.2%')

    # AMZ店铺其他费用
    vineFee = Service_Fee.loc[Service_Fee['description'].isin(['Vine'])]['total'].sum()
    Coupon_Deal_selection = PRR.loc[PRR['type'].isin([0, 'Deal Fee', 'Service Fee']) & (
        ~PRR['description'].isin(['Cost of Advertising']))]
    Coupon = PRR.loc[PRR['type'].isin([0, 'Deal Fee', 'Service Fee']) & (
        ~PRR['description'].isin(['Cost of Advertising', 'Subscription']))]
    Coupon_fee = round(Coupon['total'].sum(), 2)

    AMZ_store_other_fee = Coupon_fee + Subscription + Lightning_Deal_Fee + vineFee

    # AMZ 返还费用
    AMZ_return_fee = FBA_selling_fees_refunds + FBA_transaction_fees_refunds + Fee_Adjustment_FBA_Fee_Adjustment

    # 信用卡扣费
    AMZ_Card = round(PRR.loc[PRR['type'].isin(['Debt'])]['total'].sum(), 2)

    pt1 = skuGroup.sort_values(by='quantity', ascending=False, inplace=False)
    pt2 = refund_skuGroup.sort_values(by='quantity', ascending=False, inplace=False)
    
    pt3 = pd.DataFrame(
        {'合计销量': Order_QTY,
         '退款数量': Refund_QTY,
         'SP广告占比': spCount,
         '总销售额（1）': Order_sales,
         '亚马逊退款（2）': abs(Refunds),
         '亚马逊赔偿（3）': FBA_Inventory_Reimbursement,
         '平台处理费(平台扣点+配送费）（4）': abs(AMZnFBA),
         '仓储费（5）': abs(All_Inventory_Fee),
         '平台推广费用（6）': abs(spFee),
         '平台推广费用（信用卡扣款）': AMZ_Card,
         'AMZ店铺其他费用（7）': abs(AMZ_store_other_fee),
         'AMZ返还费用（8）': abs(AMZ_return_fee),
         }, index=[0])
    
    project_monthly_file_path = os.path.join(project_folder_path, f'{project_name}_{report_date}_monthly_{current_time}.xlsx')
    
    # 使用 ExcelWriter 的上下文管理器
    with pd.ExcelWriter(project_monthly_file_path, engine='xlsxwriter') as writer:
        # 创建一个字典来存储所有需要写入的数据帧和对应的sheet名称
        sheets_to_write = {
            '报表总览': pt3,
            '总销售额': Order,
            '销售明细': pt1,
            '亚马逊退款': Refund,
            '退货明细': pt2,
            '亚马逊赔偿': Adjustment,
            '平台处理费': AMZnFBA_Selection,
            '仓储费+移除或弃置费': FBA_Inventory_Fee,
            '亚马逊AGL物流': AGL_Selection,
            '亚马逊平台推广费用': Cost_of_Advertising_Selection,
            'AMZ店铺其他费用-订阅': Subscription_Selection,
            'AMZ店铺其他费用-优惠券&活动': Coupon_Deal_selection,
            'AMZ返还费用': Refund,
            '交易一览': PRR
        }
        
        # 使用循环一次性写入所有sheet
        for sheet_name, df in sheets_to_write.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # 重置样式
    reset_style(project_monthly_file_path)

    # 使用with语句读取文件内容
    with open(project_monthly_file_path, 'rb') as f:
        file_content = f.read()

    return file_content, f'Monthly Report{report_name}.xlsx'

@main_monthly.route('/monthly_report', methods=['GET', 'POST'])
def monthly_report():
    if request.method == 'POST':
        project_name = request.form.get('project_name')
        report_date = request.form.get('report_date')
        payment_range_report = request.files.get('payment_range_report')

        if not project_name or not report_date or not payment_range_report:
            flash('请填写所有必填项')
            return redirect(url_for('main_monthly.monthly_report'))

        if not allowed_file(payment_range_report.filename):
            flash('文件格式不正确')
            return redirect(url_for('main_monthly.monthly_report'))

        file_content, filename = process_monthly_report(project_name, report_date, payment_range_report)

        return send_file(
            io.BytesIO(file_content),
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    return render_template('dataset/monthly_report.html')

