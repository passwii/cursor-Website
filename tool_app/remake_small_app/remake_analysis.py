from datetime import datetime
import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Side, Border
from pywebio import start_server
from pywebio import session
from pywebio import config
from pywebio.input import *
from pywebio.output import *
import shutil


@config(title='项目数据分析工具')
def main():
    # 定义一个函数upload_data，用于收集和上传亚马逊运营数据分析项目的各种报表数据
    upload_data = input_group('亚马逊运营数据分析 BLF v240906', [
        # 输入项目名称，必须填写，提供了一些可选的项目名称
        input('项目名称', name='project_name', placeholder='请输入项目名称', required=True,
              datalist=['朗净', '典匠', '铨富', '天勤']),
        # 选择报表类型，必须选择一个
        select('报表类型', ['周报表', '月报表'], name='report_type', required=True),
        # 输入报表日期，必须填写，格式示例为年-月
        input('报表日期', name='report_date', placeholder='输入日期范围，月度则为2024-07', required=True),
        # 上传商品报表文件，必须上传，文件类型为.xlsx
        file_upload('商品报表-AD Product', name='ad_product_file', accept='.xlsx', required=True),
        # 上传业务报告文件，必须上传，文件类型为.csv
        file_upload('业务报告-Business Report', name='business_file', accept='.csv', required=True),
        # 上传付款报表文件，必须上传，文件类型为.csv
        file_upload('付款报表-Payment Report', name='payment_file', accept='.csv', required=True),
        # 上传基础信息表文件，必须上传，文件类型为.csv
        file_upload('基础信息表-Basic Information', name='basic_file', accept='.csv', required=True),
    ])
    # 获取当前工作目录，作为源文件夹路径
    source_folder = os.getcwd()
    # 将当前工作目录切换到源文件夹
    os.chdir(source_folder)

    # 从上传数据中提取项目相关信息
    project_name = upload_data['project_name']
    report_type = upload_data['report_type']
    report_date = upload_data['report_date']

    # 从上传数据中提取各类文件
    ad_product = upload_data['ad_product_file']
    business = upload_data['business_file']
    payment = upload_data['payment_file']
    sku_basic = upload_data['basic_file']

    # 从各类文件中提取内容
    ad_product_content = ad_product['content']
    business_content = business['content']
    payment_content = payment['content']
    sku_basic_content = sku_basic['content']

    # 根据项目名称构建项目文件夹路径
    project_folder_path = os.path.join(source_folder, 'project', project_name)

    # 创建并写入广告产品报告文件
    with open(f'{project_name}_{report_date}_AD_Product.xlsx', 'wb') as xlsx_file:
        xlsx_file.write(ad_product_content)

    # 创建并写入商业报告文件
    with open(f'{project_name}_{report_date}_Business_Report.csv', 'wb') as csv_file:
        csv_file.write(business_content)

    # 创建并写入支付报告文件
    with open(f'{project_name}_{report_date}_Payment_Report.csv', 'wb') as csv_file:
        csv_file.write(payment_content)

    # 创建并写入基本产品信息文件
    with open(f'{project_name}_Basic_Information.csv', 'wb') as csv_file:
        csv_file.write(sku_basic_content)

    # if upload_data['basic_file'] is not None:
    #     sku_basic = upload_data['basic_file']
    #     sku_basic_content = sku_basic['content']
    #     with open(f'{project_name}_Basic_Information.csv', 'wb') as csv_file:
    #         csv_file.write(sku_basic_content)
    #     basic_file_path = f'{project_name}_Basic_Information.csv'
    # else:
    #     basic_file_path = os.path.join(project_folder_path, f'{project_name}_Basic_Information.csv')

    # 读取业务报告数据
    df_business = pd.read_csv(f'{project_name}_{report_date}_Business_Report.csv', encoding='utf-8')

    # 读取支付报告数据，跳过前7行，使用','作为千位分隔符
    df_payment = pd.read_csv(f'{project_name}_{report_date}_Payment_Report.csv', encoding='utf-8', skiprows=7,
                             thousands=',')

    # 读取广告产品数据
    df_ad = pd.read_excel(f'{project_name}_{report_date}_AD_Product.xlsx', engine='openpyxl')

    # 读取基本产品信息数据
    df_sku_asin_basic = pd.read_csv(f'{project_name}_Basic_Information.csv', encoding='utf-8')

    # 根据项目文件夹路径创建必要的临时文件夹
    os.makedirs(project_folder_path, exist_ok=True)

    tmp_file_path = os.path.join(project_folder_path, 'tmp_file')
    os.makedirs(tmp_file_path, exist_ok=True)

    # 根据项目文件夹路径，生成项目总览文件的完整路径
    project_overview_file_path = os.path.join(project_folder_path, 'project_overview.xlsx')

    # 获取当前时间，并格式化为年-月-日_时-分-秒的形式，用于生成唯一文件名
    current_time = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')

    # 复制并且清理已上传的文件
    files_to_copy = [
        (f'{project_name}_{report_date}_AD_Product.xlsx', f'{tmp_file_path}/{project_name}_{report_date}_AD_Product_{current_time}.xlsx'),
        (f'{project_name}_{report_date}_Business_Report.csv', f'{tmp_file_path}/{project_name}_{report_date}_Business_Report_{current_time}.csv'),
        (f'{project_name}_{report_date}_Payment_Report.csv', f'{tmp_file_path}/{project_name}_{report_date}_Payment_Report_{current_time}.csv'),
        (f'{project_name}_Basic_Information.csv', f'{tmp_file_path}/{project_name}_Basic_Information_{current_time}.csv')
    ]

    for src_file, dest_file in files_to_copy:
        shutil.copy(src_file, dest_file)
        os.remove(src_file)

    # 复制SKU-ASIN基础信息DataFrame，用于后续项目相关的SKU-ASIN处理
    df_project_sku_asin = df_sku_asin_basic.copy()

    # 广告数据读取
    # 从广告数据DataFrame中提取'广告SKU'和'广告ASIN'两列，并创建新的DataFrame
    df_ad_sku_asin = df_ad[['广告SKU', '广告ASIN']].copy()
    # 将列名'广告SKU'和'广告ASIN'分别重命名为'SKU'和'ASIN'，以方便后续操作
    df_ad_sku_asin.rename(columns={'广告SKU': 'SKU', '广告ASIN': 'ASIN'}, inplace=True)
    # 去除重复的SKU和ASIN组合，保留唯一记录
    df_ad_sku_asin = df_ad_sku_asin.drop_duplicates()

    # 遍历广告数据中的每一行，以匹配和更新项目数据
    for index, row in df_ad_sku_asin.iterrows():
        sku = row['SKU']
        asin = row['ASIN']

        # 如果SKU在项目数据中已存在，则更新对应的ASIN
        if sku in df_project_sku_asin['SKU'].values:
            df_project_sku_asin.loc[df_project_sku_asin['SKU'] == sku, 'ASIN'] = asin
        else:
            # 如果SKU不存在，则创建并添加新行到项目数据
            new_row = pd.DataFrame({'SKU': [sku], 'ASIN': [asin]})
            df_project_sku_asin = pd.concat([df_project_sku_asin, new_row], ignore_index=True)

    # 删除重复的SKU和ASIN组合，保留第一次出现的记录
    df_project_sku_asin = df_project_sku_asin.drop_duplicates(subset=['SKU', 'ASIN'], keep='first')
    # # 将更新后的项目数据保存到文件（注释部分，可能未使用）
    # df_project_sku_asin.to_csv(project_sku_asin_file_path, index=False, encoding='utf-8')

    # 定义广告数据中需要的列，并创建一个简化的广告数据框架
    ad_column = ['广告SKU', '展示量', '点击量', '花费', '7天总销售额', '7天总销售量(#)']
    df_ad_simple = (df_ad[ad_column].copy()).groupby('广告SKU').sum().reset_index()
    # 重命名列以更清晰地表示数据含义
    df_ad_simple.rename(columns={'广告SKU': 'SKU',
                                 '展示量': '广告展示量',
                                 '点击量': '广告点击量',
                                 '花费': '广告花费',
                                 '7天总销售量(#)': '广告订单量',
                                 '7天总销售额': '广告销售额'}, inplace=True)

    # 合并简化的广告数据和项目数据，基于SKU进行连接
    df_merge_ad_sku_asin = pd.merge(df_ad_simple, df_project_sku_asin, on='SKU', how='left')

    # 读取Payment数据表，此时数据表中的sku都是小写格式
    df_payment.fillna(0, inplace=True)

    # 筛选出类型为'Order'和'Refund'的记录
    df_order_and_refund = df_payment.loc[df_payment['type'].isin(['Order', 'Refund'])]
    # 分别筛选出类型为'Order'和'Refund'的记录
    df_orders = df_payment.loc[df_payment['type'].isin(['Order'])]
    df_refund = df_payment.loc[df_payment['type'].isin(['Refund'])]

    # 计算每个sku的销售和退款的平台费用与FBA配送费
    sale_refund_amz_fee = df_order_and_refund.groupby('sku', as_index=False).agg({
        'selling fees': lambda x: round(x.mul(-1).sum(), 2),
        'fba fees': lambda x: round(x.mul(-1).sum(), 2)
    })

    # 计算每个sku的销售总额，包括产品销售、运费和促销折扣
    sale_group = df_orders.groupby('sku', as_index=False).agg({'quantity': 'sum',
                                                               'product sales': lambda x: round(x.sum(), 2),
                                                               'shipping credits': lambda x: round(x.sum(), 2),
                                                               'promotional rebates': lambda x: round(x.sum(), 2)
                                                               })

    # 计算每个sku的退款总额，包括产品销售、运费、促销折扣和其他费用
    refund_group = df_refund.groupby('sku', as_index=False).agg({'quantity': 'sum',
                                                                 'product sales': lambda x: round(x.sum(), 2),
                                                                 'shipping credits': lambda x: round(x.sum(), 2),
                                                                 'promotional rebates': lambda x: round(x.sum(), 2),
                                                                 'other': lambda x: round(x.sum(), 2)
                                                                 })

    # 计算销售数据的总销售额
    sale_group['sale total'] = (sale_group['product sales'] + sale_group['shipping credits']
                                + sale_group['promotional rebates'])

    # 计算退款数据的总退款额
    refund_group['refund total'] = abs(refund_group['product sales'] + refund_group['shipping credits']
                                       + refund_group['promotional rebates'] + refund_group['other'])

    # 重命名销售数据的列名，以便更清晰地反映数据内容
    sale_group = sale_group.rename(columns={
        'quantity': '实际销售量',
        'sale total': '实际销售额'
    })
    # 移除不再需要的列
    sale_group = sale_group.drop(columns=['product sales', 'shipping credits', 'promotional rebates'])

    # 重命名退款数据的列名，以便更清晰地反映数据内容
    refund_group = refund_group.rename(columns={
        'quantity': '实际退款量',
        'refund total': '实际退款额'
    })
    # 移除不再需要的列
    refund_group = refund_group.drop(columns=['product sales', 'shipping credits', 'promotional rebates', 'other'])

    # 合并销售和退款数据
    df_sale_refund_only = pd.merge(sale_group, refund_group, on='sku', how='left')
    # 合并销售退款数据和平台费用与FBA配送费数据
    df_sale_refund = pd.merge(df_sale_refund_only, sale_refund_amz_fee, on='sku', how='left')

    # 重命名最终数据的列名，使其更易于理解
    df_sale_refund.rename(columns={
        'sku': 'SKU',
        'selling fees': '平台佣金',
        'fba fees': 'FBA配送费'
    }, inplace=True)

    # 根据SKU将销售退货数据与项目SKU信息进行合并，以获取每个SKU的头程单价和FOB单价信息
    df_sale_refund_basic_info = pd.merge(df_sale_refund,
                                         df_project_sku_asin[['SKU', '头程单价', 'FOB单价']], on='SKU', how='left')

    # 由于已通过合并获取了所需信息，因此删除原列以简化数据集
    df_sale_refund_basic_info = df_sale_refund_basic_info.drop(columns=['头程单价', 'FOB单价'])

    # 选择保留的关键业务指标列，这些指标对于分析销售和市场表现至关重要
    business_to_keep = ['（子）ASIN', '页面浏览量 – 总计', '已订购商品数量', '会话次数 – 总计', '已订购商品销售额']
    # 从原始业务报告数据框中提取保留的列，并创建一个新的数据框
    business_report = df_business[business_to_keep].copy()

    # 重命名列，以便更清晰地反映数据的含义，简化后续分析
    business_report.rename(columns={'（子）ASIN': 'ASIN',
                                    '页面浏览量 – 总计': '页面浏览量',
                                    '已订购商品数量': '总销量',
                                    '会话次数 – 总计': '总访客',
                                    '已订购商品销售额': '总销售额',
                                    }, inplace=True)

    # 合并广告与业务数据，基于ASIN关联，以便整合不同数据源的信息
    df_merge_business_ad_asin = pd.merge(df_merge_ad_sku_asin, business_report, on='ASIN', how='left')
    # 进一步合并销售与退款基本信息，基于SKU关联，完善数据视图
    df_merge_business_ad_basic = pd.merge(df_merge_business_ad_asin, df_sale_refund_basic_info, on='SKU',
                                          how='left')
    # 创建最终的数据合并副本，为后续计算和分析做准备
    df_merge_all = df_merge_business_ad_basic.copy()
    # 计算产品的FOB成本，即实际销售量乘以FOB单价，用于分析成本
    df_merge_all['产品FOB'] = df_merge_all['实际销售量'] * df_merge_all['FOB单价']
    # 计算销售过程中的头程成本，即实际销售量乘以头程单价，用于分析物流成本
    df_merge_all['销售头程'] = df_merge_all['实际销售量'] * df_merge_all['头程单价']

    # 总览所有数据
    # 业务报告：页面浏览量	总访客	总销量	总转化	总销售额
    # 广告数据：广告展示量	广告点击	广告点击率	广告订单量	广告单占比	广告转化	广告花费	广告销售额	CPC	ACOS	SP
    # 自然数据；自然流量	自然流量占比	自然单	自然单转化	退款量	退款率
    # Payment数据：产品FOB	成本占比	销售头程	头程占比	配送费	配送费占比	佣金	佣金占比	总成本	总成本占比	利润	利润率	实际销售额	平均售价
    overview_data = {
        '日期': [],
        # 基础信息
        'SKU': [],
        'ASIN': [],
        # 业务报告
        '总销量': [],
        '总销售额': [],
        '页面浏览量': [],
        '总访客': [],
        '总转化': [],  # 总转化 = 总销量 / 页面浏览量
        # 广告数据
        '广告展示量': [],
        '广告点击量': [],
        '广告点击率': [],  # 广告点击率 = 广告点击量 / 广告展示量
        '广告订单量': [],
        '广告单占比': [],  # 广告单占比 = 广告订单量 / 实际销售量
        '广告转化率': [],  # 广告转化 = 广告订单量 / 广告点击量
        '广告花费': [],
        '广告销售额': [],
        'CPC': [],  # CPC = 广告花费 / 广告点击量
        'ACOS': [],  # ACOS = 广告花费 / 广告销售额
        'SP占比': [],  # SP占比 = 广告花费 / 总销售额
        # 自然数据
        '自然流量': [],  # 自然流量 = 页面浏览量 - 广告点击量
        '自然流量占比': [],  # 自然流量占比 = 自然流量 / 页面浏览量
        '自然单': [],  # 自然单 = 总销量 - 广告订单量
        '自然单转化': [],  # 自然单转化 = 自然单 / 自然流量

        # Payment数据
        '产品FOB': [],
        '成本占比': [],
        '销售头程': [],
        '头程占比': [],
        'FBA配送费': [],
        '配送费占比': [],
        '平台佣金': [],
        '佣金占比': [],
        '总成本': [],
        '总成本占比': [],
        '利润': [],
        '利润率': [],
        '实际销售额': [],
        '实际销售量': [],
        '实际退款额': [],
        '实际退款量': [],
        '平均售价': [],  # 平均售价 = 实际销售额 / 实际销售量
        '退款率': [],  # 退款率 = 实际退款 / 实际销量'
    }

    # 汇总
    df_overview = pd.DataFrame(overview_data)
    df_overview = pd.concat([df_overview, df_merge_all], ignore_index=True)
    # df_overview = df_overview.drop_duplicates(subset=['SKU', 'ASIN'], keep='first')
    df_overview['日期'] = report_date

    # 处理前的NA填充
    df_overview['实际销售额'] = df_overview['实际销售额'].fillna(0)
    df_overview['实际退款额'] = df_overview['实际退款额'].fillna(0)
    df_overview['实际销售量'] = df_overview['实际销售量'].fillna(0)

    # 业务报告
    df_overview['总销量'] = df_overview['总销量'].fillna(0)
    df_overview['总销售额'] = df_overview['总销售额'].fillna(0).str.replace(',', '').str.replace('US$', '').astype(
        float)
    df_overview['页面浏览量'] = pd.to_numeric(df_overview['页面浏览量'].fillna(0).astype(str).str.replace(',', ''),
                                              errors='coerce')
    df_overview['总访客'] = pd.to_numeric(df_overview['总访客'].fillna(0).astype(str).str.replace(',', ''),
                                          errors='coerce')
    df_overview['总访客'] = pd.to_numeric(df_overview['总访客'].fillna(0).astype(str).str.replace(',', ''),
                                          errors='coerce')

    # 计算总转化率，当页面浏览量为0时，总转化为0
    df_overview['总转化'] = np.where(df_overview['页面浏览量'] == 0, 0,
                                     df_overview['总销量'] / df_overview['页面浏览量']).round(4)

    # 自然流量
    df_overview['自然流量'] = df_overview['页面浏览量'] - df_overview['广告点击量']
    df_overview['自然流量占比'] = (df_overview['自然流量'] / df_overview['页面浏览量']).round(4)
    df_overview['自然单'] = df_overview['总销量'] - df_overview['广告订单量']
    df_overview['自然单转化'] = np.where(df_overview['自然流量'] == 0, 0,
                                         df_overview['自然单'] / df_overview['自然流量'])

    # 广告
    df_overview['广告花费'] = df_overview['广告花费'].round(2)
    df_overview['广告销售额'] = df_overview['广告销售额'].round(2)
    df_overview['广告展示量'] = df_overview['广告展示量'].astype(int)
    df_overview['广告点击量'] = df_overview['广告点击量'].astype(int)
    df_overview['广告订单量'] = df_overview['广告订单量'].astype(int)
    df_overview['广告点击率'] = np.where(df_overview['广告展示量'] == 0, 0,
                                         df_overview['广告点击量'] / df_overview['广告展示量']).round(4)
    df_overview['广告单占比'] = np.where(df_overview['实际销售量'] == 0, 0,
                                         df_overview['广告订单量'] / df_overview['实际销售量']).round(4)
    df_overview['广告转化率'] = (df_overview['广告订单量'] / df_overview['广告点击量']).round(4)
    df_overview['CPC'] = np.where(df_overview['广告点击量'] == 0, 0,
                                  df_overview['广告花费'] / df_overview['广告点击量']).round(2)
    df_overview['ACOS'] = np.where(df_overview['广告销售额'] == 0, 0,
                                   df_overview['广告花费'] / df_overview['广告销售额'])
    df_overview['SP占比'] = np.where(df_overview['广告花费'] == 0, 0,
                                     df_overview['广告花费'] / df_overview['实际销售额']).round(4)

    # Payment数据
    df_overview['实际销售额'] = df_overview['实际销售额'].round(2)
    df_overview['实际退款额'] = df_overview['实际退款额'].round(2)
    df_overview['实际退款量'] = pd.to_numeric(df_overview['实际退款量'].fillna(0), errors='coerce').astype(int)

    df_overview['产品FOB'] = df_overview['产品FOB'].fillna(0)
    df_overview['成本占比'] = np.where(df_overview['实际销售额'] == 0, 0,
                                       df_overview['产品FOB'] / df_overview['实际销售额']).round(4)
    df_overview['销售头程'] = df_overview['销售头程'].fillna(0)
    df_overview['头程占比'] = np.where(df_overview['实际销售额'] == 0, 0,
                                       df_overview['销售头程'] / df_overview['实际销售额']).round(4)
    df_overview['FBA配送费'] = df_overview['FBA配送费'].fillna(0)
    df_overview['配送费占比'] = np.where(df_overview['实际销售额'] == 0, 0,
                                         df_overview['FBA配送费'] / df_overview['实际销售额']).round(4)

    df_overview['总成本'] = df_overview['产品FOB'] + df_overview['销售头程'] + df_overview['FBA配送费'] + df_overview[
        '平台佣金']
    df_overview['总成本占比'] = np.where(df_overview['实际销售额'] == 0, 0,
                                         df_overview['总成本'] / df_overview['实际销售额']).round(4)
    df_overview['利润'] = (df_overview['实际销售额'] - df_overview['总成本'] - df_overview['广告花费'] - df_overview[
        '实际退款额']).round(2)
    df_overview['利润率'] = np.where(df_overview['实际销售额'] == 0, 0,
                                     df_overview['利润'] / df_overview['实际销售额']).round(4)

    df_overview['平均售价'] = np.where(df_overview['实际销售量'] == 0, 0,
                                       df_overview['实际销售额'] / df_overview['实际销售量']).round(2)
    df_overview['退款率'] = np.where(df_overview['实际销售额'] == 0, 0,
                                     df_overview['实际退款额'] / df_overview['实际销售额']).round(4)
    df_overview['佣金占比'] = np.where(df_overview['实际销售量'] == 0, 0,
                                       df_overview['平台佣金'] / df_overview['实际销售额']).round(4)

    # 定义不需要的列名列表，因为这些列可能包含敏感信息或不再需要
    overview_drop_cols = ['头程单价', 'FOB单价']
    # 从概览数据框中删除指定的列，以净化数据或准备数据共享
    df_overview = df_overview.drop(columns=overview_drop_cols)

    # 指定项目概览模板文件的路径，为了加载模板以便填充数据或进行其他操作
    template_file_path = 'project/project_overview_template.xlsx'
    # 加载Excel工作簿，以便可以编辑或操作数据
    wb = load_workbook(template_file_path)
    # 获取当前活动的工作表，准备对其进行操作
    ws = wb.active

    # 定义边框样式
    center_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(border_style='thin', color='FF000000'),
                         right=Side(border_style='thin', color='FF000000'),
                         top=Side(border_style='thin', color='FF000000'),
                         bottom=Side(border_style='thin', color='FF000000'))

    # 遍历数据框架中的每一行每一列，并设置单元格格式
    for r_idx, row in enumerate(dataframe_to_rows(df_overview, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = center_alignment  # 设置单元格内容居中
            cell.border = thin_border  # 设置单元格边框为细边框
            # 如果标题行的单元格格式为百分比，则对该单元格也应用相同的百分比格式
            if ws.cell(row=1, column=c_idx).number_format == '0.00%':
                cell.number_format = '0.00%'

    # 保存项目概览工作簿到指定路径
    wb.save(project_overview_file_path)

    # 构造新的项目概览文件名，包含项目名称、概览类型和报告日期
    new_project_ov_name = f'{project_name}_overview_{report_date}.xlsx'

    # 拼接新的项目概览文件的完整路径
    new_project_ov_path = os.path.join(project_folder_path, new_project_ov_name)

    # 如果新项目概览路径已存在，则删除它，以确保最新的项目概览可以正确重命名和放置
    if os.path.exists(new_project_ov_path):
        os.remove(new_project_ov_path)
    # 重命名项目概览文件到新的路径，这反映了项目概览的更新
    os.rename(project_overview_file_path, new_project_ov_path)

    # 这里使用rb（二进制读）模式读取文件内容，是因为put_file()函数可能需要原始字节数据
    file_content = open(new_project_ov_path, 'rb').read()
    # 调用put_file()函数更新项目概览，这是更新项目信息的一个关键步骤
    session.download(name=new_project_ov_name, content=file_content)

    # 调用session的hold方法，暂停会话活动
    session.hold()


start_server(main, host='0.0.0.0', port=8200, debug=True, auto_open_webbrowser=False, remote_access=False,
             reconnect_timeout=60)
