from flask import Blueprint, render_template, request, send_file, redirect, flash, url_for
import os
import io
import datetime
import pandas as pd
import numpy as np
from openpyxl.utils.dataframe import dataframe_to_rows
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, NamedStyle
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter

main_product_analysis = Blueprint('main_product_analysis', __name__)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'csv', 'xlsx'}

def save_file(file, folder, filename):
    if not os.path.exists(folder):
        os.makedirs(folder)
    filepath = os.path.join(folder, filename)
    file.save(filepath)
    return filepath

def process_product_analysis(project_name, report_start_date, report_end_date, business_report, payment_report, ad_product_report, basic_report):
    current_time = datetime.datetime.now().strftime('%H-%M-%S')
    source_folder = os.getcwd()
    os.chdir(source_folder)

    # Combine dates for file naming
    start_date = report_start_date.replace('-', '')
    end_date = report_end_date.replace('-', '')[4:]  # Remove the year (YYYY) from end date
    report_date = f"{start_date}-{end_date}"

    project_folder_path = os.path.join(source_folder, 'project', project_name, '产品数据分析')
    os.makedirs(project_folder_path, exist_ok=True)

    tmp_folder_path = os.path.join(source_folder, 'project', project_name, 'tmp')
    os.makedirs(tmp_folder_path, exist_ok=True)

    product_analysis_file_path = os.path.join(project_folder_path, f'{project_name}_product_analysis_{report_date}.xlsx')

    business_report_path = f'{project_name}_Business_Report_{report_date}.csv'
    payment_report_path = f'{project_name}_Payment_Report_{report_date}.csv'
    ad_product_report_path = f'{project_name}_AD_Product_{report_date}.xlsx'
    basic_report_path = f'{project_name}_Basic_Information.csv'

    with open(business_report_path, 'wb') as f:
        f.write(business_report.read())

    with open(payment_report_path, 'wb') as f:
        f.write(payment_report.read())

    with open(ad_product_report_path, 'wb') as f:
        f.write(ad_product_report.read())

    with open(basic_report_path, 'wb') as f:
        f.write(basic_report.read())

    business_report = pd.read_csv(business_report_path, encoding='utf-8')
    payment_report = pd.read_csv(payment_report_path, encoding='utf-8', thousands=',', skiprows=7)
    ad_product_report = pd.read_excel(ad_product_report_path, engine='openpyxl')
    basic_report = pd.read_csv(basic_report_path, encoding='utf-8')


    files_to_copy = [
        (business_report_path, f'{tmp_folder_path}/{project_name}_Business_Report_{report_date}_{current_time}.csv'),
        (payment_report_path, f'{tmp_folder_path}/{project_name}_Payment_Report_{report_date}_{current_time}.csv'),
        
        (ad_product_report_path, f'{tmp_folder_path}/{project_name}_AD_Product_{report_date}_{current_time}.xlsx'),
        (basic_report_path, f'{tmp_folder_path}/{project_name}_Basic_Information_{report_date}_{current_time}.csv')
    ]

    for src, dst in files_to_copy:
        shutil.copy(src, dst)
        os.remove(src)

    # 复制SKU-ASIN基础信息DataFrame，用于后续项目相关的SKU-ASIN处理
    df_project_sku_asin = basic_report.copy()

    # 广告数据读取
    df_ad_sku_asin = ad_product_report[['广告SKU', '广告ASIN']].copy()
    df_ad_sku_asin.rename(columns={'广告SKU': 'SKU', '广告ASIN': 'ASIN'}, inplace=True)
    df_ad_sku_asin = df_ad_sku_asin.drop_duplicates()

    # 遍历广告数据中的每一行，以匹配和更新项目数据
    for index, row in df_ad_sku_asin.iterrows():
        sku = row['SKU']
        asin = row['ASIN']

        if sku in df_project_sku_asin['SKU'].values:
            df_project_sku_asin.loc[df_project_sku_asin['SKU'] == sku, 'ASIN'] = asin
        else:
            new_row = pd.DataFrame({'SKU': [sku], 'ASIN': [asin]})
            df_project_sku_asin = pd.concat([df_project_sku_asin, new_row], ignore_index=True)
    
    # 删除重复的SKU和ASIN组合，保留第一次出现的记录
    df_project_sku_asin = df_project_sku_asin.drop_duplicates(subset=['SKU', 'ASIN'], keep='first')

    ad_column = ['广告SKU', '展示量', '点击量', '花费', '7天总销售额', '7天总销售量(#)']
    df_ad_simple = (ad_product_report[ad_column].copy()).groupby('广告SKU').sum().reset_index()
    df_ad_simple.rename(columns={'广告SKU': 'SKU',
                                 '展示量': '广告展示量',
                                 '点击量': '广告点击量',
                                 '花费': '广告花费',
                                 '7天总销售量(#)': '广告订单量',
                                 '7天总销售额': '广告销售额'}, inplace=True)
    # 将广告数据与项目数据进行合并, 让 SKU-ASIN 表加入到广告数据表中
    df_merge_ad_sku_asin = pd.merge(df_ad_simple, df_project_sku_asin, on='SKU', how='left')

    # 读取Payment数据表，此时数据表中的sku都是小写格式
    df_payment = payment_report.copy()
    df_payment.fillna(0, inplace=True)


    # 筛选出类型为'Order'和'Refund'的记录
    df_order_and_refund = df_payment.loc[df_payment['type'].isin(['Order', 'Refund'])]
    # 分别筛选出类型为'Order'和'Refund'的记录
    df_orders = df_payment.loc[df_payment['type'].isin(['Order'])]
    df_refund = df_payment.loc[df_payment['type'].isin(['Refund'])]

    # 计算每个sku的销售和退款的平台费用与FBA配送费
    sale_refund_amz_fee = df_order_and_refund.groupby('sku', as_index=False).agg({
        'selling fees': lambda x: round(x.mul(-1).sum(), 2),
        'fba fees': lambda x: round(x.mul(-1).sum(), 2)
    })

    # 计算每个sku的销售总额，包括产品销售、运费和促销折扣
    sale_group = df_orders.groupby('sku', as_index=False).agg({'quantity': 'sum',
                                                               'product sales': lambda x: round(x.sum(), 2),
                                                               'shipping credits': lambda x: round(x.sum(), 2),
                                                               'promotional rebates': lambda x: round(x.sum(), 2)
    })

    # 计算每个sku的退款总额，包括产品销售、运费、促销折扣和其他费用
    df_refund_group = df_refund.groupby('sku', as_index=False).agg({'quantity': 'sum',
                                                                   'product sales': lambda x: round(x.sum(), 2),
                                                                   'shipping credits': lambda x: round(x.sum(), 2),
                                                                   'promotional rebates': lambda x: round(x.sum(), 2),
                                                                   'other': lambda x: round(x.sum(), 2)
    })

    # 计算销售数据的总销售额
    sale_group['sale total'] = (sale_group['product sales'] + sale_group['shipping credits']
                                + sale_group['promotional rebates'])

    # 计算退款数据的总退款额
    df_refund_group['refund total'] = abs(df_refund_group['product sales'] + df_refund_group['shipping credits']
                                          + df_refund_group['promotional rebates'] + df_refund_group['other'])  
    
    # 重命名销售数据的列名，以便更清晰地反映数据内容
    sale_group = sale_group.rename(columns={
        'quantity': '实际销售量',
        'sale total': '实际销售额'
    })
    # 移除不再需要的列
    sale_group = sale_group.drop(columns=['product sales', 'shipping credits', 'promotional rebates'])
    
    # 重命名退款数据的列名，以便更清晰地反映数据内容
    df_refund_group = df_refund_group.rename(columns={
        'quantity': '实际退款量',
        'refund total': '实际退款额'
    })
    # 移除不再需要的列
    df_refund_group = df_refund_group.drop(columns=['product sales', 'shipping credits', 'promotional rebates', 'other'])   
    
    # 合并销售和退款数据
    df_sale_refund_only = pd.merge(sale_group, df_refund_group, on='sku', how='left')
    # 合并销售退款数据和平台费用与FBA配送费数据
    df_sale_refund = pd.merge(df_sale_refund_only, sale_refund_amz_fee, on='sku', how='left')
    
    # 重命名最终数据的列名，使其更易于理解
    df_sale_refund.rename(columns={
        'sku': 'SKU',
        'selling fees': '平台佣金',
        'fba fees': 'FBA配送费'
    }, inplace=True)
    
    # 根据SKU将销售退款数据与项目SKU信息进行合并，以获取每个SKU的头程单价和FOB单价信息
    df_sale_refund_basic_info = pd.merge(df_sale_refund,
                                         df_project_sku_asin[['SKU', '头程单价', 'FOB单价']], on='SKU', how='left')
    
    # 由于已通过合并获取了所需信息，因此删除原列以简化数据集
    df_sale_refund_basic_info = df_sale_refund_basic_info.drop(columns=['头程单价', 'FOB单价'])
    
    # 选择保留的关键业务指标列，这些指标对于分析销售和市场表现至关重要
    business_to_keep = ['（子）ASIN', '页面浏览量 - 总计 ', '已订购商品数量', '会话数 - 总计', '已订购商品销售额']
    # 从原始业务报告数据框中提取保留的列，并创建一个新的数据框
    business_report = business_report[business_to_keep].copy()
    
    # 重命名列，以便更清晰地反映数据的含义，简化后续分析
    business_report.rename(columns={'（子）ASIN': 'ASIN',
                                    '页面浏览量 - 总计 ': '页面浏览量',
                                    '已订购商品数量': '总销量',
                                    '会话数 - 总计': '总访客',
                                    '已订购商品销售额': '总销售额',
                                    }, inplace=True)
    # 将总销售额列，逗号和 US$符号去除,并且数值化
    business_report['总销售额'] = business_report['总销售额'].replace({'\$': '', ',': '', 'US': ''}, regex=True).astype(float)

    # 合并广告与业务数据，基于ASIN关联，以便整合不同数据源的信息
    df_merge_business_ad_asin = pd.merge(df_merge_ad_sku_asin, business_report, on='ASIN', how='left')
    # 进一步合并销售与退款基本信息，基于SKU关联，完善数据视图
    df_merge_business_ad_basic = pd.merge(df_merge_business_ad_asin, df_sale_refund_basic_info, on='SKU',
                                          how='left')
    # 创建最终的数据合并副本，为后续计算和分析做准备
    df_merge_all = df_merge_business_ad_basic.copy()
    # 计算产品的FOB成本，即实际销售量乘以FOB单价，用于分析成本
    df_merge_all['产品FOB'] = df_merge_all['实际销售量'] * df_merge_all['FOB单价']
    
    # 计算销售过程中的头程成本，即实际销售量乘以头程单价，用于分析物流成本
    df_merge_all['销售头程'] = df_merge_all['实际销售量'] * df_merge_all['头程单价']
    
    # 总览所有数据
    # 业务报告：页面浏览量	总访客	总销量	总转化	总销售额
    # 广告数据：广告展示量	广告点击	广告点击率	广告订单量	广告单占比	广告转化	广告花费	广告销售额	CPC	ACOS	SP
    # 自然数据；自然流量	自然流量占比	自然单	自然单转化	退款量	退款率
    # Payment数据：产品FOB	成本占比	销售头程	头程占比	配送费	配送费占比	佣金	佣金占比	总成本	总成本占比	利润	利润率	实际销售额	平均售价
    
    overview_data = {
        '日期': [],
        # 基础信息
        'SKU': [],
        'ASIN': [],
        # 业务报告
        '总销量': [],
        '总销售额': [],
        '页面浏览量': [],
        '总访客': [],
        '总转化': [],  # 总转化 = 总销量 / 页面浏览量
        # 广告数据
        '广告展示量': [],
        '广告点击量': [],
        '广告点击率': [],  # 广告点击率 = 广告点击量 / 广告展示量
        '广告订单量': [],
        '广告单占比': [],  # 广告单占比 = 广告订单量 / 业务报告销售量
        '广告转化率': [],  # 广告转化 = 广告订单量 / 广告点击量
        '广告花费': [],
        '广告销售额': [],
        'CPC': [],  # CPC = 广告花费 / 广告点击量
        'ACOS': [],  # ACOS = 广告花费 / 广告销售额
        'SP占比': [],  # SP占比 = 广告花费 / 总销售额
        # 自然数据
        '自然流量': [],  # 自然流量 = 页面浏览量 - 广告点击量
        '自然流量占比': [],  # 自然流量占比 = 自然流量 / 页面浏览量
        '自然单': [],  # 自然单 = 总销量 - 广告订单量
        '自然单转化': [],  # 自然单转化 = 自然单 / 自然流量

        # Payment数据
        '产品FOB': [],
        '成本占比': [],
        '销售头程': [],
        '头程占比': [],
        'FBA配送费': [],
        '配送费占比': [],
        '平台佣金': [],
        '佣金占比': [],
        '总成本': [],
        '总成本占比': [],
        '利润': [],
        '利润率': [],
        '实际销售额': [],
        '实际销售量': [],
        '实际退款额': [],
        '实际退款量': [],
        '平均售价': [],  # 平均售价 = 实际销售额 / 实际销售量
        '退款率': [],  # 退款率 = 实际退款 / 实际销量'
    }

    # 汇总
    df_overview = pd.DataFrame(overview_data)
    df_overview = pd.concat([df_overview, df_merge_all], ignore_index=True)
    # df_overview = df_overview.drop_duplicates(subset=['SKU', 'ASIN'], keep='first')
    df_overview['日期'] = report_date

    # 处理前的NA填充
    df_overview['实际销售额'] = df_overview['实际销售额'].fillna(0)
    df_overview['实际退款额'] = df_overview['实际退款额'].fillna(0)
    df_overview['实际销售量'] = df_overview['实际销售量'].fillna(0)

    # 业务报告
    df_overview['总销量'] = df_overview['总销量'].fillna(0)
    df_overview['总销售额'] = df_overview['总销售额'].fillna(0)
    df_overview['页面浏览量'] = pd.to_numeric(df_overview['页面浏览量'].fillna(0).astype(str).str.replace(',', ''), errors='coerce')
    df_overview['总访客'] = pd.to_numeric(df_overview['总访客'].fillna(0).astype(str).str.replace(',', ''), errors='coerce')
    
    # 计算总转化率，当页面浏览量为0时，总转化为0
    df_overview['总转化'] = np.where(df_overview['页面浏览量'] == 0, 0, df_overview['总销量'] / df_overview['页面浏览量'])
    
    # 自然流量
    df_overview['自然流量'] = df_overview['页面浏览量'] - df_overview['广告点击量']
    df_overview['自然流量占比'] = (df_overview['自然流量'] / df_overview['页面浏览量']).round(4)
    df_overview['自然单'] = df_overview['总销量'] - df_overview['广告订单量']
    df_overview['自然单转化'] = np.where(df_overview['自然流量'] == 0, 0, df_overview['自然单'] / df_overview['自然流量'])
    
    # 广告
    df_overview['广告花费'] = df_overview['广告花费'].round(2)
    df_overview['广告销售额'] = df_overview['广告销售额'].round(2)
    df_overview['广告展示量'] = df_overview['广告展示量'].astype(int)
    df_overview['广告点击量'] = df_overview['广告点击量'].astype(int)
    df_overview['广告订单量'] = df_overview['广告订单量'].astype(int)
    df_overview['广告点击率'] = np.where(df_overview['广告展示量'] == 0, 0, df_overview['广告点击量'] / df_overview['广告展示量'])
    df_overview['广告单占比'] = np.where(df_overview['总销量'] == 0, 0, df_overview['广告订单量'] / df_overview['总销量'])
    df_overview['广告转化率'] = df_overview['广告订单量'] / df_overview['广告点击量']
    df_overview['CPC'] = np.where(df_overview['广告点击量'] == 0, 0, df_overview['广告花费'] / df_overview['广告点击量']).round(2)
    df_overview['ACOS'] = np.where(df_overview['广告销售额'] == 0, 0, df_overview['广告花费'] / df_overview['广告销售额'])
    df_overview['SP占比'] = np.where(df_overview['总销售额'] == 0, 0, df_overview['广告花费'] / df_overview['总销售额'])
    
    # Payment数据
    df_overview['实际销售额'] = df_overview['实际销售额'].round(2)
    df_overview['实际退款额'] = df_overview['实际退款额'].round(2)
    df_overview['实际退款量'] = pd.to_numeric(df_overview['实际退款量'].fillna(0), errors='coerce').astype(int)
    
    df_overview['产品FOB'] = df_overview['产品FOB'].fillna(0)
    df_overview['成本占比'] = np.where(df_overview['实际销售额'] == 0, 0, df_overview['产品FOB'] / df_overview['实际销售额'])
    df_overview['销售头程'] = df_overview['销售头程'].fillna(0)
    df_overview['头程占比'] = np.where(df_overview['实际销售额'] == 0, 0, df_overview['销售头程'] / df_overview['实际销售额'])
    df_overview['FBA配送费'] = df_overview['FBA配送费'].fillna(0)
    df_overview['配送费占比'] = np.where(df_overview['实际销售额'] == 0, 0, df_overview['FBA配送费'] / df_overview['实际销售额'])
    df_overview['平台佣金'] = df_overview['平台佣金'].fillna(0)
    df_overview['佣金占比'] = np.where(df_overview['实际销售量'] == 0, 0, df_overview['平台佣金'] / df_overview['实际销售额'])
    df_overview['总成本'] = df_overview['产品FOB'] + df_overview['销售头程'] + df_overview['FBA配送费'] + df_overview['平台佣金']
    df_overview['总成本占比'] = np.where(df_overview['实际销售额'] == 0, 0, df_overview['总成本'] / df_overview['实际销售额'])
    df_overview['利润'] = df_overview['实际销售额'] - df_overview['总成本'] - df_overview['广告花费'] - df_overview['实际退款额']
    df_overview['利润率'] = np.where(df_overview['实际销售额'] == 0, 0, df_overview['利润'] / df_overview['实际销售额'])
    
    df_overview['平均售价'] = np.where(df_overview['实际销售量'] == 0, 0, (df_overview['实际销售额'] / df_overview['实际销售量']).round(2))
    df_overview['退款率'] = np.where(df_overview['实际销售额'] == 0, 0, df_overview['实际退款额'] / df_overview['实际销售额'])
    
    overview_drop_cols = ['头程单价', 'FOB单价']
    df_overview = df_overview.drop(columns=overview_drop_cols)

    # 指定项目概览模板文件的路径，为了加载模板以便填充数据或进行其他操作
    template_file_path = 'model_file/product_analysis_template.xlsx'
    # 加载Excel工作簿，以便可以编辑或操作数据
    wb = load_workbook(template_file_path)
    # 获取当前活动的工作表，准备对其进行操作
    ws = wb.active
    
    # 定义边框样式
    center_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(border_style='thin', color='FF000000'),
                         right=Side(border_style='thin', color='FF000000'),
                         top=Side(border_style='thin', color='FF000000'),
                         bottom=Side(border_style='thin', color='FF000000'))

    # 遍历数据框架中的每一行每一列，并设置单元格格式
    for r_idx, row in enumerate(dataframe_to_rows(df_overview, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = center_alignment  # 设置单元格内容居中
            cell.border = thin_border  # 设置单元格边框为细边框
            # 如果标题行的单元格格式为百分比，则对该单元格也应用相同的百分比格式
            if ws.cell(row=1, column=c_idx).number_format == '0.00%':
                cell.number_format = '0.00%'

    # 保存项目概览工作簿到指定路径
    wb.save(product_analysis_file_path)
    
    with open(product_analysis_file_path, 'rb') as f:
        file_content = f.read()

    return file_content, f'{project_name}_product_analysis_{report_date}.xlsx'

@main_product_analysis.route('/product_analysis', methods=['GET', 'POST'])
def product_analysis():
    if request.method == 'POST':
        project_name = request.form.get('project_name')
        report_start_date = request.form.get('report_start_date')
        report_end_date = request.form.get('report_end_date')
        business_report = request.files.get('business_report')
        payment_report = request.files.get('payment_report')
        ad_product_report = request.files.get('ad_product_report')
        basic_report = request.files.get('basic_report')
        
        if not (project_name and report_start_date and report_end_date and business_report and payment_report and ad_product_report and basic_report):
            flash('请填写所有字段并上传所有文件')
            return redirect(url_for('main_product_analysis.product_analysis'))
        
        if not (allowed_file(business_report.filename) and allowed_file(payment_report.filename) and allowed_file(ad_product_report.filename) and allowed_file(basic_report.filename)):
            flash('文件格式不正确')
            return redirect(url_for('main_product_analysis.product_analysis'))
        
        file_content, filename = process_product_analysis(project_name, report_start_date, report_end_date, business_report, payment_report, ad_product_report, basic_report)  
        
        return send_file(
            io.BytesIO(file_content),
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    return render_template('dataset/product_analysis.html')
